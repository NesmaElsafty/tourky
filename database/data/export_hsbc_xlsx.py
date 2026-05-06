"""One-off export: HSBC xlsx -> hsbc_routes.json (UTF-8)."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "hsbc-routes.xlsx"
OUT = ROOT / "hsbc_routes.json"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    routes: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        points: list[dict] = []
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            d = dict(zip(header, row, strict=False))

            def cell_str(v: object) -> str | None:
                if v is None:
                    return None
                if hasattr(v, "isoformat"):
                    return str(v)
                return str(v).strip() or None

            lat = d.get("Latitude")
            lon = d.get("Longitude")
            if lat is None or lon is None:
                continue
            points.append(
                {
                    "node": cell_str(d.get("Name")),
                    "lat": float(lat),
                    "lon": float(lon),
                    "neighborhood_en": cell_str(d.get("English Neighborhood")),
                    "poi_en": cell_str(d.get("English Primary PoI")),
                }
            )
        routes.append({"sheet": sheet_name, "points": points})

    OUT.write_text(
        json.dumps(routes, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {OUT} ({len(routes)} routes)")


if __name__ == "__main__":
    sys.exit(main())
